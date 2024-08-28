import requests
from typing import Text
import requests
import json
import datetime,time
from datetime import timedelta, datetime
import openpyxl

def func_b():
    try:
        api_key='eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwODAxdjEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczODM2MDA0MSwiaWQiOiJhZWY4N2E1MS0wODRkLTRkNjYtOTA0ZS02MjNhOTUzODVmOTciLCJpaWQiOjI0NTE0NTA3LCJvaWQiOjg3OTYzNSwicyI6MTA0LCJzaWQiOiJhZmNhMTg2NC1mOWY4LTQ1MDYtOTM3Yy0wMzZlN2E1YTUwM2EiLCJ0IjpmYWxzZSwidWlkIjoyNDUxNDUwN30.pK0miWktv4dmaMk3nRgET_0HYWqhef2_Y_hMSFl6s379LGGyYlJCVWHhQIfycQXQPHRcSh31kTSrEzhnmryK8A'


        id_dict = {}
        url = 'https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter' 
        headers = {"Authorization": api_key}
        r = requests.get(url, headers = headers, params={"limit":1000})
        for row in r.json()['data']['listGoods']:
            id_dict[row['nmID']] = row['vendorCode']


        id_dict_2 = {}
        path_new = "Справочник номенклатур.xlsx"
        wb_new = openpyxl.load_workbook(path_new)  
        sheet_new = wb_new['Справочник номенклатуры']
        for row in sheet_new.values:
            id_dict_2[row[1]] = f'{row[6]}{row[9]}{row[10]}'


        #*******Списки кампаний**********
        url = 'https://advert-api.wildberries.ru/adv/v1/promotion/count' 
        headers = {"Authorization": api_key}
        r = requests.get(url, headers = headers, json={
        })

        path = "Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet = wb['списки кампаний']

        with open('data.json', 'w') as file:
            json.dump(r.json(), file)
        num = 3
        for i in range(len(r.json()['adverts'])):
            data = r.json()['adverts'][i]['advert_list']          
            for row in data:
                sheet[f'A{num}'] = row['advertId']        
                sheet[f'B{num}'] = datetime.strptime(row['changeTime'][:10], '%Y-%m-%d').strftime('%d.%m.%Y')
                status_cam = {-1: 'кампания в процессе удаления', 4: 'готова к запуску', 7: 'кампания завершена', 8: 'отказался', 9: 'идут показы', 11: 'кампания на паузе'}
                type_cam = {4: 'кампания в каталоге', 5: 'кампания в карточке товара', 6: 'кампания в поиске', 7: 'кампания в рекомендациях на главной странице', 8: 'автоматическая кампания', 9: 'поиск + каталог'}
                try:
                    sheet[f'C{num}'] = type_cam[r.json()['adverts'][i]['type']]
                except:
                    sheet[f'C{num}'] = r.json()['adverts'][i]['type']
                try:
                    sheet[f'D{num}'] = status_cam[r.json()['adverts'][i]['status']]
                except:
                    sheet[f'D{num}'] = r.json()['adverts'][i]['type']
                num += 1

        wb.save("Общая модель WB.xlsx")



        #***********Информация о кампаниях******
        url = 'https://advert-api.wildberries.ru/adv/v1/promotion/adverts' 
        headers = {"Authorization": api_key}
        l = []
        with open('data.json', 'r', encoding='utf-8') as f:
            text = json.load(f) 
            for i in range(len(text['adverts'])):
                data = text['adverts'][i]['advert_list']        
                for row in data:
                    l.append(row['advertId'])

        path = "Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet = wb['информация кампании']
        num = 1
        counter = 0
        while counter != len(l):
            if len(l) - counter > 50:
                r = requests.post(url, headers = headers, json=l[counter:counter+50])
                counter += 50
            else:
                r = requests.post(url, headers = headers, json=l[counter:])
                counter += len(l) - counter 
            for row in r.json():
                sheet[f'A{num}'] = datetime.strptime(row['endTime'][:10], '%Y-%m-%d').strftime('%d.%m.%Y')
                sheet[f'B{num}'] = datetime.strptime(row['createTime'][:10], '%Y-%m-%d').strftime('%d.%m.%Y')
                sheet[f'C{num}'] = datetime.strptime(row['changeTime'][:10], '%Y-%m-%d').strftime('%d.%m.%Y')
                sheet[f'D{num}'] = datetime.strptime(row['startTime'][:10], '%Y-%m-%d').strftime('%d.%m.%Y')
                sheet[f'E{num}'] = row['name']
                sheet[f'F{num}'] = row['dailyBudget']
                sheet[f'G{num}'] = row['advertId']
                sheet[f'H{num}'] = row['status']
                sheet[f'I{num}'] = row['type']
                sheet[f'J{num}'] = row['paymentType']
                try:
                    sheet[f'K{num}'] = row['searchPluseState']
                except:
                    pass  
                try:
                    sheet[f'X{num}'] = row['autoParams']['subject']['id']
                    sheet[f'Y{num}'] = row['autoParams']['subject']['name']
                except:
                    pass  
                try:
                    for row_1 in row['autoParams']['sets']:
                        try:    
                            data = sheet[f'Z{num}'].value 
                            if data:    
                                sheet[f'Z{num}'] = f"{data}, {row_1['id']}"
                                data = sheet[f'AA{num}'].value             
                                sheet[f'AA{num}'] = f"{data}, {row_1['name']}"
                            else:
                                sheet[f'Z{num}'] = row_1['id']
                                data = sheet[f'AA{num}'].value             
                                sheet[f'AA{num}'] = row_1['name']
                        except:
                            pass  
                except:
                    pass
                try:
                    for row_1 in row['autoParams']['menus']:                
                        try:
                            data = sheet[f'AB{num}'].value  
                            if data: 
                                sheet[f'AB{num}'] = f"{data}, {row_1['id']}"
                                data = sheet[f'AC{num}'].value  
                                sheet[f'AC{num}'] = f"{data}, {row_1['name']}"
                            else:
                                sheet[f'AB{num}'] = row_1['id']
                                sheet[f'AC{num}'] = row_1['name']
                        except:
                            pass  
                except:
                    pass
                try:
                    sheet[f'AD{num}'] = row['autoParams']['active']['carousel']
                    sheet[f'AE{num}'] = row['autoParams']['active']['recom']
                    sheet[f'AF{num}'] = row['autoParams']['active']['booster']
                except:
                    pass    
                try:
                    for row_1 in row['autoParams']['nmCPM']:                       
                        try:
                            data = sheet[f'AG{num}'].value  
                            if data: 
                                sheet[f'AG{num}'] = f"{data}, {row_1['nm']}"
                                data = sheet[f'AH{num}'].value  
                                sheet[f'AH{num}'] = f"{data}, {row_1['cpm']}"
                            else:
                                sheet[f'AG{num}'] = row_1['nm']
                                sheet[f'AH{num}'] = row_1['cpm']
                        except:
                            pass  
                except:
                    pass
                try:
                    for row_1 in row['autoParams']['nms']:            
                        try:
                            data = sheet[f'AI{num}'].value  
                            if data: 
                                sheet[f'AI{num}'] = f"{data}, {row_1}"
                            else:
                                sheet[f'AI{num}'] = row_1
                        except:
                            pass  
                except:
                    pass
                try:
                    for row_1 in row['autoParams']['cpm']:            
                        try:
                            data = sheet[f'AJ{num}'].value  
                            if data: 
                                sheet[f'AJ{num}'] = f"{data}, {row_1}"
                            else:
                                sheet[f'AJ{num}'] = row_1
                        except:
                            pass  
                except:
                    pass
                try:
                    sheet[f'AL{num}'] = row['unitedParams'][0]['subject']['id']
                    sheet[f'AM{num}'] = row['unitedParams'][0]['subject']['name']
                except:
                    pass 
                try:
                    for row_1 in row['unitedParams'][0]['menus']:                       
                        try:
                            data = sheet[f'AN{num}'].value  
                            if data: 
                                sheet[f'AN{num}'] = f"{data}, {row_1['id']}"
                                data = sheet[f'AN{num}'].value  
                                sheet[f'AO{num}'] = f"{data}, {row_1['name']}"
                            else:
                                sheet[f'AN{num}'] = row_1['id']
                                sheet[f'AO{num}'] = row_1['name']
                        except:
                            pass  
                except:
                    pass
                try:
                    for row_1 in row['unitedParams'][0]['nms']:            
                        try:
                            data = sheet[f'AP{num}'].value  
                            if data: 
                                sheet[f'AP{num}'] = f"{data}, {row_1}"
                            else:
                                sheet[f'AP{num}'] = row_1
                        except:
                            pass  
                except:
                    pass  
                try:
                    sheet[f'AK{num}'] = row['unitedParams'][0]['searchCPM']            
                except:
                    pass   
                try:
                    sheet[f'AQ{num}'] = row['unitedParams'][0]['catalogCPM']           
                except:
                    pass  
                num += 1  
        wb.save("Общая модель WB.xlsx")


        #******Статистика кампаний******
        api_key='eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwODAxdjEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczODM2MDA0MSwiaWQiOiJhZWY4N2E1MS0wODRkLTRkNjYtOTA0ZS02MjNhOTUzODVmOTciLCJpaWQiOjI0NTE0NTA3LCJvaWQiOjg3OTYzNSwicyI6MTA0LCJzaWQiOiJhZmNhMTg2NC1mOWY4LTQ1MDYtOTM3Yy0wMzZlN2E1YTUwM2EiLCJ0IjpmYWxzZSwidWlkIjoyNDUxNDUwN30.pK0miWktv4dmaMk3nRgET_0HYWqhef2_Y_hMSFl6s379LGGyYlJCVWHhQIfycQXQPHRcSh31kTSrEzhnmryK8A'
        url = 'https://advert-api.wildberries.ru/adv/v2/fullstats'
        headers = {"Authorization": api_key}

        path = "Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet3 = wb['информация кампании']

        l = []

        with open('data.json', 'r', encoding='utf-8') as f:
            text = json.load(f) 
            for i in range(len(text['adverts'])):
                data = text['adverts'][i]['advert_list']
                
                for row in data:
                    l.append({'id':row['advertId'], 'dates': [(datetime.now()- timedelta(days = 1)).strftime('%Y-%m-%d') ]}) 
                    

        sheet = wb['статистика кампаний']
        num = 3#len([row for row in sheet.values if row[0]]) + 1
        counter = 0

        while counter <= len(l):    
            if len(l) - counter >= 100:
                r = requests.post(url, headers = headers, json=l[counter:counter + 100])                        
                counter += 100
            else:
                r = requests.post(url, headers = headers, json=l[counter:])
                counter += len(l) - counter

            sheet_2 = wb['информация кампании']
            list_method = {}
            for row in sheet_2.values:
                list_method[row[6]] = row[4]

            try:    
                id_comp = r.json()[0]['advertId']
                for row in r.json()[0]['days'][0]['apps']:
                    try:
                        if row['appType'] == 1:
                            appType = 'сайт'
                        if row['appType'] == 32:
                            appType = 'android'
                        else:
                            appType = 'ios'        
                        for row_2 in row['nm']:
                            sheet[f'A{num}'] = (datetime.now() - timedelta(days=1)).strftime('%d.%m.%Y')
                            sheet[f'B{num}'] = id_comp  

                            for row in sheet3.values:
                                if str(row[6]) == str(id_comp):         
                                    sheet[f'C{num}'] = row[4]
                                    fff = {-1: 'кампания в процессе удаления',                                 
                                            4: 'готова к запуску',
                                            7: 'кампания завершена',
                                            8: 'отказался',
                                            9: 'идут показы',
                                            11: 'кампания на паузе'}
                                    sheet[f'H{num}'] = fff[row[7]]
                                    fff = {4: 'кампания в каталоге (устаревший тип)',
                                            5: 'кампания в карточке товара (устаревший тип)',
                                            6: 'кампания в поиске (устаревший тип)',
                                            7: 'кампания в рекомендациях на главной странице (устаревший тип)',
                                            8: 'автоматическая кампания',
                                            9: 'Аукцион'}
                                    sheet[f'I{num}'] = fff[row[8]]
                                    sheet[f'J{num}'] = row[10]
                                    sheet[f'K{num}'] = row[13]
                                    sheet[f'L{num}'] = row[14]
                                    sheet[f'M{num}'] = row[22]
                                    sheet[f'N{num}'] = row[29]
                                    sheet[f'O{num}'] = row[30]
                                    sheet[f'P{num}'] = row[31]
                                    break

                            sheet[f'D{num}'] = row_2['name'] #название товара
                            sheet[f'E{num}'] = row_2['nmId']
                            try:
                                sheet[f'F{num}'] = id_dict[row_2['nmId']]
                                #print(row_2['nmId'], id_dict[row_2['nmId']], id_dict_2[id_dict[row_2['nmId']]])
                            except:
                                pass
                            try:
                                sheet[f'G{num}'] = id_dict_2[id_dict[row_2['nmId']]]
                            except:
                                pass

                            sheet[f'R{num}'] = row_2['views']
                            sheet[f'S{num}'] = row_2['clicks']
                            sheet[f'T{num}'] = row_2['ctr']
                            sheet[f'U{num}'] = row_2['cpc']
                            sheet[f'V{num}'] = row_2['sum']
                            sheet[f'W{num}'] = row_2['atbs']
                            sheet[f'X{num}'] = row_2['orders']
                            sheet[f'Y{num}'] = row_2['cr']
                            sheet[f'Z{num}'] = row_2['shks']
                            sheet[f'AA{num}'] = row_2['sum_price']
                            sheet[f'AB{num}'] = appType
                            num += 1
                    except:
                        pass

                                
                
            except:
                pass
            time.sleep(60)  
            wb.save("Общая модель WB.xlsx")
                

    except:
        file = open('Ошибки.txt', 'a')
        file.write(f"{datetime.now().strftime('%d.%m.%Y %HS%M:%S')} Ошибка, вероятно нет подключения к интернету\n")

func_b()
