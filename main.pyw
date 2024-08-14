from aiogram import Bot, Dispatcher, executor, types
import config as cfg
import markups as nav
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.contrib.fsm_storage.memory import MemoryStorage
import requests
import os
from dotenv import load_dotenv
import json
from datetime import timedelta, datetime
import openpyxl
import asyncio
from func_uch import *

load_dotenv()

bot = Bot(token=cfg.TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())



class States(StatesGroup):
    step_1 = State()
    step_2 = State()
    step_3 = State()
    step_4 = State()
    step_5 = State()
    step_6 = State()
    step_7 = State()
    step_8 = State()
    step_9 = State()
    step_10 = State()

async def my_func():
    while True:
        if datetime.now().hour == 0 and datetime.now().minute == 50:
            func_b()
            await asyncio.sleep(3000)  
        await asyncio.sleep(5) 

async def on_startup(dp):
    asyncio.create_task(my_func())

@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    if message.from_user.id in [1193989520, 424915104, 6400135693, 336194148]: #  список id пользователей, которым разрешено запускать бота
        await bot.send_message(message.from_user.id, 'Меню', reply_markup=nav.keyb_1)

@dp.callback_query_handler(lambda c: True)
async def call_back_q(callback: types.CallbackQuery):   
    if callback.data == 'btn_1':        
        try:
            await bot.delete_message(callback.message.chat.id, callback.message.message_id)
        except:
            pass  

        path = "Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet = wb['создать автоматическую кампанию']
        for row in sheet.values[1:]:
            url = 'https://advert-api.wildberries.ru/adv/v1/save-ad' #Создать автоматическую кампанию
            api_key='eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwODAxdjEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczODM2MDA0MSwiaWQiOiJhZWY4N2E1MS0wODRkLTRkNjYtOTA0ZS02MjNhOTUzODVmOTciLCJpaWQiOjI0NTE0NTA3LCJvaWQiOjg3OTYzNSwicyI6MTA0LCJzaWQiOiJhZmNhMTg2NC1mOWY4LTQ1MDYtOTM3Yy0wMzZlN2E1YTUwM2EiLCJ0IjpmYWxzZSwidWlkIjoyNDUxNDUwN30.pK0miWktv4dmaMk3nRgET_0HYWqhef2_Y_hMSFl6s379LGGyYlJCVWHhQIfycQXQPHRcSh31kTSrEzhnmryK8A'
            headers = {"Authorization": api_key}

            try:
                r = requests.post(url, headers = headers, json={
                "type": 8,
                "name": row[0],
                "subjectId": row[1],
                "sum": row[2],
                "btype": row[3],
                "on_pause": row[4],
                "nms": [
                row[5]
                ],
                "cpm": row[6]})
            except:
                await bot.send_message(callback.from_user.id, f'Ошибка в добавлении {row[0]} {row[5]}')
        await bot.send_message(callback.from_user.id, 'Меню', reply_markup=nav.keyb_1)

    elif callback.data == 'btn_2':        
        try:
            await bot.delete_message(callback.message.chat.id, callback.message.message_id)
        except:
            pass  

        path = "Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet = wb['создать кампанию аукцион']
        for row in sheet.values[1:]:
            url = 'https://advert-api.wildberries.ru/adv/v2/seacat/save-ad' #Создать кампанию Поиск + Каталог (теперь аукцион)
            api_key='eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwODAxdjEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczODM2MDA0MSwiaWQiOiJhZWY4N2E1MS0wODRkLTRkNjYtOTA0ZS02MjNhOTUzODVmOTciLCJpaWQiOjI0NTE0NTA3LCJvaWQiOjg3OTYzNSwicyI6MTA0LCJzaWQiOiJhZmNhMTg2NC1mOWY4LTQ1MDYtOTM3Yy0wMzZlN2E1YTUwM2EiLCJ0IjpmYWxzZSwidWlkIjoyNDUxNDUwN30.pK0miWktv4dmaMk3nRgET_0HYWqhef2_Y_hMSFl6s379LGGyYlJCVWHhQIfycQXQPHRcSh31kTSrEzhnmryK8A'
            headers = {"Authorization": api_key}

            try:
                r = requests.post(url, headers = headers, json={
                "name": row[0],
                "nms": [
                row[1]
                ]
                })
            except:
                await bot.send_message(callback.from_user.id, f'Ошибка в добавлении {row[0]} {row[1]}')
        await bot.send_message(callback.from_user.id, 'Меню', reply_markup=nav.keyb_1)


    elif callback.data == 'btn_3':        
        try:
            await bot.delete_message(callback.message.chat.id, callback.message.message_id)
        except:
            pass  

        path = "Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet = wb['списки кампаний']
        for row in sheet.values[1:]:
            api_key='eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwODAxdjEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczODM2MDA0MSwiaWQiOiJhZWY4N2E1MS0wODRkLTRkNjYtOTA0ZS02MjNhOTUzODVmOTciLCJpaWQiOjI0NTE0NTA3LCJvaWQiOjg3OTYzNSwicyI6MTA0LCJzaWQiOiJhZmNhMTg2NC1mOWY4LTQ1MDYtOTM3Yy0wMzZlN2E1YTUwM2EiLCJ0IjpmYWxzZSwidWlkIjoyNDUxNDUwN30.pK0miWktv4dmaMk3nRgET_0HYWqhef2_Y_hMSFl6s379LGGyYlJCVWHhQIfycQXQPHRcSh31kTSrEzhnmryK8A'
            if row[8] == 1:
                try:
                    url = 'https://advert-api.wildberries.ru/adv/v0/delete' #Удалить кампанию
                    headers = {"Authorization": api_key}
                    r = requests.get(url, headers = headers, json={"id":row[0]})
                except:
                    await bot.send_message(callback.from_user.id, f'Ошибка в удалении {row[0]}')
        await bot.send_message(callback.from_user.id, 'Меню', reply_markup=nav.keyb_1)


    elif callback.data == 'btn_4':        
        try:
            await bot.delete_message(callback.message.chat.id, callback.message.message_id)
        except:
            pass  

        path = "Общая модель.xlsx"
        wb = openpyxl.load_workbook(path)  
        sheet = wb['списки кампаний']
        for row in sheet.values[1:]:
            api_key='eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwODAxdjEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczODM2MDA0MSwiaWQiOiJhZWY4N2E1MS0wODRkLTRkNjYtOTA0ZS02MjNhOTUzODVmOTciLCJpaWQiOjI0NTE0NTA3LCJvaWQiOjg3OTYzNSwicyI6MTA0LCJzaWQiOiJhZmNhMTg2NC1mOWY4LTQ1MDYtOTM3Yy0wMzZlN2E1YTUwM2EiLCJ0IjpmYWxzZSwidWlkIjoyNDUxNDUwN30.pK0miWktv4dmaMk3nRgET_0HYWqhef2_Y_hMSFl6s379LGGyYlJCVWHhQIfycQXQPHRcSh31kTSrEzhnmryK8A'
            if row[4]:
                try:
                    url = 'https://advert-api.wildberries.ru/adv/v0/cpm' #Изменение ставки у кампании
                    headers = {"Authorization": api_key}
                    r = requests.post(url, headers = headers, json={
                    "advertId": row[0],
                    "type": int(row[5]),
                    "cpm": row[4],
                    "param": row[7],
                    "instrument": row[6]
                    })
                except:
                    await bot.send_message(callback.from_user.id, f'Ошибка в изменении {row[0]}')
        await bot.send_message(callback.from_user.id, 'Меню', reply_markup=nav.keyb_1)

if __name__ == "__main__":  
    try:  
        executor.start_polling(dp, skip_updates=True)
    except:
        file = open('Ошибки.txt', 'a')
        file.write(f"{datetime.now().strftime('%d.%m.%Y %H:%M:%S')} Ошибка, вероятно нет подключения к интернету\n")  